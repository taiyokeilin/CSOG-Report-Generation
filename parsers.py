try:
    import polars as pl
except ImportError:
    # Thin pandas shim so the module loads; Streamlit env will have polars
    import pandas as _pd
    class _PolarsFacade:
        """Minimal polars shim backed by pandas for environments without polars."""
        @staticmethod
        def read_csv(source, **kwargs):
            skip = kwargs.pop("skip_rows", 0)
            null_vals = kwargs.pop("null_values", [""])
            infer = kwargs.pop("infer_schema_length", None)
            import io as _io
            if hasattr(source, "read"):
                content = source.read()
                source = _io.StringIO(content.decode() if isinstance(content, bytes) else content)
            df = _pd.read_csv(source, skiprows=skip, na_values=null_vals)
            return _PolarsFacade._wrap(df)

        @staticmethod
        def _wrap(df):
            class _DF:
                def __init__(self, pdf): self._df = pdf
                def filter(self, expr): return _DF(self._df)
                def to_dicts(self): return self._df.to_dict("records")
                @property
                def columns(self): return list(self._df.columns)
                def __len__(self): return len(self._df)
                def head(self, n=5): return _DF(self._df.head(n))
                def to_pandas(self): return self._df
                def n_unique(self): return self._df.nunique().max()
                class _Col:
                    def __init__(self, s): self._s = s
                    def is_not_null(self): return None
                    def is_in(self, vals): return None
                    def drop_nulls(self): return self
                    def to_list(self): return self._s.dropna().tolist()
                    def unique(self): return self
                    def sort(self): return self
                def __getitem__(self, col): return _DF._DF._Col(self._df[col]) if col in self._df else _DF._Col(_pd.Series([]))
            return _DF(df)

        @staticmethod
        def DataFrame(records):
            return _pd.DataFrame(records)

        class col:
            def __init__(self, name): self.name = name
            def is_not_null(self): return None
            def is_in(self, vals): return None

    pl = _PolarsFacade()

import re
import io
from typing import Union


COMMON_SCHEMA = [
    "club", "shot_num_session", "shot_num_club",
    "ball_speed_mph", "launch_angle_deg", "side_angle_deg",
    "backspin_rpm", "side_spin_rpm", "tilt_angle_deg", "total_spin_rpm",
    "carry_yd", "total_yd", "offline_yd", "descent_angle_deg",
    "peak_height_ft", "to_pin_ft",
    "club_speed_mph", "smash_factor", "angle_of_attack_deg",
    "club_path_deg", "face_angle_deg", "face_to_path_deg",
    "dynamic_lie_deg", "dynamic_loft_deg", "closure_rate_dps",
    "face_impact_horizontal_mm", "face_impact_vertical_mm",
    "face_impact_from_center_mm", "date",
]

PIN_DISTANCES = list(range(20, 140, 10))  # 20..130


def _to_pin_ft(total_yd: float, offline_yd: float, target_yd: int) -> float | None:
    if total_yd is None or offline_yd is None or target_yd is None:
        return None
    return ((offline_yd ** 2) + (target_yd - total_yd) ** 2) ** 0.5 * 3


def _extract_target_from_name(club_name: str) -> int | None:
    """Extract leading distance number from club name, e.g. '60 yard 58*' -> 60."""
    if not club_name:
        return None
    m = re.match(r"^(\d+)", club_name.strip())
    if m:
        val = int(m.group(1))
        if val in PIN_DISTANCES:
            return val
    return None


def _parse_direction_value(val: str, pos_dir: str = "R") -> float | None:
    """Parse strings like '5.2R', '3.1L', '5.2 In-Out', '3.1 Out-In'."""
    if val is None:
        return None
    val = str(val).strip()
    try:
        return float(val)
    except ValueError:
        pass
    m = re.match(r"^([0-9]+\.?[0-9]*)\s*(.+)$", val)
    if not m:
        return None
    num, direction = float(m.group(1)), m.group(2).strip().upper()
    positive_dirs = {pos_dir.upper(), "R", "IN-OUT", "UP", "TOE UP", "HIGH", "OPEN"}
    if direction in positive_dirs:
        return num
    return -num


def _decode(file_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "windows-1252"):
        try:
            return file_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode("latin-1", errors="replace")


def parse_trackman(file_bytes: bytes) -> pl.DataFrame:
    content = _decode(file_bytes)
    df = pl.read_csv(
        io.StringIO(content),
        infer_schema_length=1000,
        null_values=["", "N/A"],
    ).filter(pl.col("Club Name").is_not_null())

    def to_pin(row_total, row_offline, row_club):
        target = _extract_target_from_name(str(row_club) if row_club else "")
        return _to_pin_ft(row_total, row_offline, target) if target else None

    rows = df.to_dicts()
    records = []
    club_counters: dict[str, int] = {}
    for i, row in enumerate(rows):
        club = row.get("Club Name")
        club_counters[club] = club_counters.get(club, 0) + 1
        total = row.get("Total Distance (yds)")
        offline = row.get("Offline (yds L-/R+)")
        peak_yd = row.get("Peak Height (yds)")
        club_path = row.get("Club Path (deg out-in-/in-out+)")
        face_target = row.get("Face to Target (deg closed-/open+)")
        face_h = row.get("Face Impact Horizontal (mm toe-/heel+)")
        face_v = row.get("Face Impact Vertical (mm low-/high+)")
        target = _extract_target_from_name(club or "")
        records.append({
            "club": club,
            "shot_num_session": i + 1,
            "shot_num_club": club_counters[club],
            "ball_speed_mph": _safe_float(row.get("Ball Speed (mph)")),
            "launch_angle_deg": _safe_float(row.get("Launch Angle (deg)")),
            "side_angle_deg": None,
            "backspin_rpm": _safe_float(row.get("Back Spin (rpm)")),
            "side_spin_rpm": _safe_float(row.get("Side Spin (rpm L-/R+)")),
            "tilt_angle_deg": None,
            "total_spin_rpm": _safe_float(row.get("Total Spin (rpm)")),
            "carry_yd": _safe_float(row.get("Carry (yds)")),
            "total_yd": _safe_float(total),
            "offline_yd": _safe_float(offline),
            "descent_angle_deg": _safe_float(row.get("Descent Angle (deg)")),
            "peak_height_ft": _safe_float(peak_yd) * 3 if _safe_float(peak_yd) is not None else None,
            "to_pin_ft": _to_pin_ft(_safe_float(total), _safe_float(offline), target),
            "club_speed_mph": _safe_float(row.get("Club Speed (mph)")),
            "smash_factor": _safe_float(row.get("Efficiency")),
            "angle_of_attack_deg": _safe_float(row.get("Angle of Attack (deg)")),
            "club_path_deg": _safe_float(club_path),
            "face_angle_deg": _safe_float(face_target),
            "face_to_path_deg": (
                (_safe_float(club_path) or 0) + (_safe_float(face_target) or 0)
                if club_path is not None and face_target is not None else None
            ),
            "dynamic_lie_deg": _safe_float(row.get("Lie (deg toe down-/toe up+)")),
            "dynamic_loft_deg": _safe_float(row.get("Loft (deg)")),
            "closure_rate_dps": _safe_float(row.get("Closure Rate (deg/sec)")),
            "face_impact_horizontal_mm": _safe_float(face_h),
            "face_impact_vertical_mm": _safe_float(face_v),
            "face_impact_from_center_mm": (
                ((_safe_float(face_h) or 0) ** 2 + (_safe_float(face_v) or 0) ** 2) ** 0.5
                if face_h is not None and face_v is not None else None
            ),
            "date": str(row.get("Shot Created Date", "")),
        })
    return pl.DataFrame(records)


def parse_foresight(file_bytes: bytes) -> pl.DataFrame:
    content = _decode(file_bytes)
    df = pl.read_csv(
        io.StringIO(content),
        skip_rows=1,
        infer_schema_length=1000,
        null_values=["", "N/A"],
    ).filter(
        pl.col("Club").is_not_null() &
        ~pl.col("Club").is_in(["Average", "Std. Dev."])
    )

    rows = df.to_dicts()
    records = []
    club_counters: dict[str, int] = {}
    for i, row in enumerate(rows):
        club = row.get("Club")
        club_counters[club] = club_counters.get(club, 0) + 1
        total = _safe_float(row.get("Total"))
        offline_raw = row.get("Offline")
        offline = _parse_direction_value(str(offline_raw)) if offline_raw else None
        carry = _safe_float(row.get("Carry"))
        peak_yd = _safe_float(row.get("Peak Height"))
        face_h_raw = row.get("Face Impact Lateral")
        face_v_raw = row.get("Face Impact Vertical")
        face_h = _parse_direction_value(str(face_h_raw), "Toe") if face_h_raw else None
        face_v = _parse_direction_value(str(face_v_raw), "High") if face_v_raw else None
        aoa_raw = row.get("Angle of Attack")
        cp_raw = row.get("Club Path")
        ftp_raw = row.get("Face to Path")
        lie_raw = row.get("Lie")
        side_raw = row.get("Side Angle")
        sidespin_raw = row.get("Sidespin")
        tilt_raw = row.get("Tilt Angle")
        target = _extract_target_from_name(club or "")
        club_path = _parse_direction_value(str(cp_raw), "In-Out") if cp_raw else None
        ftp = _parse_direction_value(str(ftp_raw), "Open") if ftp_raw else None
        records.append({
            "club": club,
            "shot_num_session": i + 1,
            "shot_num_club": club_counters[club],
            "ball_speed_mph": _safe_float(row.get("Ball Speed")),
            "launch_angle_deg": _safe_float(row.get("Launch Angle")),
            "side_angle_deg": _parse_direction_value(str(side_raw)) if side_raw else None,
            "backspin_rpm": _safe_float(row.get("Backspin")),
            "side_spin_rpm": _parse_direction_value(str(sidespin_raw)) if sidespin_raw else None,
            "tilt_angle_deg": _parse_direction_value(str(tilt_raw)) if tilt_raw else None,
            "total_spin_rpm": _safe_float(row.get("Total Spin")),
            "carry_yd": carry,
            "total_yd": total,
            "offline_yd": offline,
            "descent_angle_deg": _safe_float(row.get("Descent Angle")),
            "peak_height_ft": peak_yd * 3 if peak_yd is not None else None,
            "to_pin_ft": _to_pin_ft(total, offline, target),
            "club_speed_mph": _safe_float(row.get("Club Speed")),
            "smash_factor": _safe_float(row.get("Efficiency")),
            "angle_of_attack_deg": _parse_direction_value(str(aoa_raw), "Up") if aoa_raw else None,
            "club_path_deg": club_path,
            "face_angle_deg": (ftp - club_path) if (ftp is not None and club_path is not None) else None,
            "face_to_path_deg": ftp,
            "dynamic_lie_deg": _parse_direction_value(str(lie_raw), "Toe Up") if lie_raw else None,
            "dynamic_loft_deg": _safe_float(row.get("Loft")),
            "closure_rate_dps": _safe_float(row.get("Closure Rate")),
            "face_impact_horizontal_mm": face_h,
            "face_impact_vertical_mm": face_v,
            "face_impact_from_center_mm": (
                ((face_h or 0) ** 2 + (face_v or 0) ** 2) ** 0.5
                if face_h is not None and face_v is not None else None
            ),
            "date": None,
        })
    return pl.DataFrame(records)


def parse_flightscope(file_bytes: bytes) -> pl.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row_vals, name):
        try:
            idx = headers.index(name)
            return row_vals[idx]
        except (ValueError, IndexError):
            return None

    rows_data = list(ws.iter_rows(min_row=2, values_only=True))
    records = []
    club_counters: dict[str, int] = {}
    session_num = 0
    for row in rows_data:
        club = col(row, "club")
        shot = col(row, "Shot")
        if not club or shot in ("Avg", "Dev"):
            continue
        session_num += 1
        club_counters[club] = club_counters.get(club, 0) + 1
        ball_mph = _safe_float(col(row, "Ball (mph)"))
        club_mph = _safe_float(col(row, "Club (mph)"))
        smash = round(ball_mph / club_mph, 2) if ball_mph and club_mph else None
        total = _safe_float(col(row, "Total (yds)"))
        lat_raw = col(row, "Lateral (yds)")
        offline = _parse_direction_value(str(lat_raw)) if lat_raw else None
        ftp_raw = col(row, "FTP (°)")
        ftt_raw = col(row, "FTT (°)")
        cp_raw = col(row, "Club Path (°)")
        face_h_in = _safe_float(col(row, "Lateral Impact (in)"))
        face_v_in = _safe_float(col(row, "Vertical Impact (in)"))
        face_h = face_h_in * 25.4 if face_h_in is not None else None
        face_v = face_v_in * 25.4 if face_v_in is not None else None
        target = _extract_target_from_name(club or "")
        records.append({
            "club": club,
            "shot_num_session": session_num,
            "shot_num_club": club_counters[club],
            "ball_speed_mph": ball_mph,
            "launch_angle_deg": _safe_float(col(row, "Launch V (°)")),
            "side_angle_deg": None,
            "backspin_rpm": None,
            "side_spin_rpm": None,
            "tilt_angle_deg": None,
            "total_spin_rpm": _safe_float(col(row, "Spin (rpm)")),
            "carry_yd": _safe_float(col(row, "Carry (yds)")),
            "total_yd": total,
            "offline_yd": offline,
            "descent_angle_deg": _safe_float(col(row, "DescentV (°)")),
            "peak_height_ft": _safe_float(col(row, "Height (ft)")),
            "to_pin_ft": _to_pin_ft(total, offline, target),
            "club_speed_mph": club_mph,
            "smash_factor": smash,
            "angle_of_attack_deg": _safe_float(col(row, "AOA (°)")),
            "club_path_deg": _parse_direction_value(str(cp_raw)) if cp_raw else None,
            "face_angle_deg": _parse_direction_value(str(ftt_raw)) if ftt_raw else None,
            "face_to_path_deg": _parse_direction_value(str(ftp_raw)) if ftp_raw else None,
            "dynamic_lie_deg": None,
            "dynamic_loft_deg": _safe_float(col(row, "Dynamic Loft (°)")),
            "closure_rate_dps": None,
            "face_impact_horizontal_mm": face_h,
            "face_impact_vertical_mm": face_v,
            "face_impact_from_center_mm": (
                ((face_h or 0) ** 2 + (face_v or 0) ** 2) ** 0.5
                if face_h is not None and face_v is not None else None
            ),
            "date": None,
        })
    return pl.DataFrame(records)


def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_file(file_bytes: bytes, monitor_type: str) -> pl.DataFrame:
    t = monitor_type.lower()
    if t == "trackman":
        return parse_trackman(file_bytes)
    elif t == "foresight":
        return parse_foresight(file_bytes)
    elif t == "flightscope":
        return parse_flightscope(file_bytes)
    raise ValueError(f"Unknown monitor type: {monitor_type}")
