import io
import math
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from calculations import (
    compute_club_stats, compute_overall,
    feet_to_feet_inches_str, compute_proximity_ft,
    get_target_proximity_ft, get_target_range_yd, get_target_rate,
)
from data.tour_targets import get_tour_target, get_level_multipliers

# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER_BG   = "FFFFFF"
C_HEADER_FG   = "000000"
C_SECTION_BG  = "FFFFFF"
C_SECTION_FG  = "000000"
C_COL_HDR_BG  = "F2F2F2"
C_DIST_BG     = "FFFFFF"
C_ALT_ROW     = "FFFFFF"
C_GREEN       = "E2EFDA"
C_AMBER       = "FFEB9C"
C_RED_BG      = "FFC7CE"
C_RED_FG      = "9C0006"
C_GREEN_FG    = "375623"
C_AMBER_FG    = "7D6608"
C_RAW_FG      = "000000"

THIN = Side(style="thin", color="AAAAAA")
MED  = Side(style="medium", color="555555")
FULL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _border(left=None, right=None, top=None, bottom=None):
    return Border(left=left or Side(), right=right or Side(),
                  top=top or Side(), bottom=bottom or Side())

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=15, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_row(ws, row, values: list, bold=False, bg=None, fg="000000",
             border=None, align="left", size=15):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _font(bold=bold, color=fg, size=size)
        if bg:
            cell.fill = _fill(bg)
        if border:
            cell.border = border
        cell.alignment = _center() if align == "center" else _left()


def _merge_title(ws, row, text, start_col, end_col, bg, fg, size=14, bold=True, align="center"):
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row, end_column=end_col
    )
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = _font(bold=bold, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = _center() if align == "center" else _left()
    return cell


def _col_header_row(ws, row, headers: list, ncols: int):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, italic=True, size=13, name="Arial")
        cell.fill = _fill(C_COL_HDR_BG)
        cell.alignment = _center()
        cell.border = FULL_BORDER


def _goal_status_style(cell, status):
    if status == "Goal Met":
        cell.fill = _fill(C_GREEN)
        cell.font = _font(color=C_GREEN_FG, size=14)
    elif status == "Approaching Goal":
        cell.fill = _fill(C_AMBER)
        cell.font = _font(color=C_AMBER_FG, size=14)
    elif status == "Goal in Progress":
        cell.fill = _fill(C_RED_BG)
        cell.font = _font(color=C_RED_FG, size=14)
    else:
        cell.font = _font(size=14)



def _pct_style(cell, actual_pct, target_pct, size=14):
    """Color a percentage cell using the same green/amber/red logic as goal status."""
    if actual_pct is None or target_pct is None or target_pct == 0:
        cell.font = _font(size=size)
        return
    ratio = actual_pct / target_pct
    if ratio >= 1.0:
        cell.fill = _fill(C_GREEN)
        cell.font = _font(color=C_GREEN_FG, size=size)
    elif ratio >= 0.7:
        cell.fill = _fill(C_AMBER)
        cell.font = _font(color=C_AMBER_FG, size=size)
    else:
        cell.fill = _fill(C_RED_BG)
        cell.font = _font(color=C_RED_FG, size=size)

# ── Excel formula helpers ──────────────────────────────────────────────────
def _formula_target_proximity(dist_cell: str, level: int) -> str:
    """
    Build an Excel formula that, given a distance cell reference, looks up
    proximity target using embedded lookup table logic.
    We embed a nested IFS over the known distances rather than needing a
    hidden sheet, so editing distance recalculates instantly.
    """
    cases = []
    from data.tour_targets import TOUR_TARGETS, LEVELS_MULTIPLIERS
    _, prox_mult = LEVELS_MULTIPLIERS.get(level, (None, None))
    for yd, (prox_ft, range_yd, rate) in sorted(TOUR_TARGETS.items()):
        rounded_val = f'ROUND({dist_cell}/10,0)*10'
        cases.append(f'ROUND({dist_cell}/10,0)*10={yd},{prox_ft * prox_mult}')
    ifs_body = ",".join(cases)
    return f'=IFERROR(IFS({ifs_body}),"")'


def _formula_target_range(dist_cell: str, level: int) -> str:
    from data.tour_targets import TOUR_TARGETS, LEVELS_MULTIPLIERS
    _, prox_mult = LEVELS_MULTIPLIERS.get(level, (None, None))
    cases = []
    for yd, (prox_ft, range_yd, rate) in sorted(TOUR_TARGETS.items()):
        if range_yd is not None:
            cases.append(f'ROUND({dist_cell}/10,0)*10={yd},{range_yd * prox_mult}')
    ifs_body = ",".join(cases)
    return f'=IFERROR(IFS({ifs_body}),"")'


def _formula_target_rate(dist_cell: str, level: int) -> str:
    from data.tour_targets import TOUR_TARGETS, LEVELS_MULTIPLIERS
    rate_mult, _ = LEVELS_MULTIPLIERS.get(level, (None, None))
    cases = []
    for yd, (prox_ft, range_yd, rate) in sorted(TOUR_TARGETS.items()):
        if rate is not None:
            scaled = min(1.0, rate * rate_mult)
            cases.append(f'ROUND({dist_cell}/10,0)*10={yd},{scaled}')
    ifs_body = ",".join(cases)
    return f'=IFERROR(IFS({ifs_body}),"")'


def build_report_sheet(
    ws,
    session_info: dict,
    sections: list[dict],
    raw_data_sheet_name: str,
    overall: dict,
    logo_path: str = None,
):
    NCOLS = 12
    COL_CLUB   = 1
    COL_LEVEL  = 2
    COL_TTYPE  = 3
    COL_DIST   = 4   # ← yellow, editable
    COL_TRAW   = 5
    COL_ARAW   = 6
    COL_TPCT   = 7
    COL_ATT    = 8
    COL_SUCC   = 9
    COL_APCT   = 10
    COL_GOAL   = 11
    COL_STATUS = 12

    # Column widths
    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 17
    ws.column_dimensions["F"].width = 17
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 0
    ws.column_dimensions["N"].width = 0

    current_row = 1

    # ── Row heights ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25

    # ── Title spans all columns ───────────────────────────────────────────────
    _merge_title(ws, current_row, "Player Practice Report Card",
                 1, NCOLS, C_HEADER_BG, C_HEADER_FG, size=18)
    current_row += 1

    # ── Logo: floating image anchored A1, sized as square filling rows 1-3 ───
    # Row 1=45pt, Row 2=25pt, Row 3=25pt → total ~95pt ≈ 127px (at 96dpi)
    # Col A=15 units, B=9.5 units, C=18 units → ~42.5 units * 7px ≈ 298px wide
    # Use square = min of height/width available to keep it square
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            logo_size = 160   # points → fits snugly in rows 1-3 height
            img.width = logo_size
            img.height = logo_size
            img.anchor = "C1"
            ws.add_image(img)
        except Exception:
            pass

    # ── Row 2: Date | Player: | Player Name | Coach: | Coach Name ────────────
    meta_row = current_row
    for c in range(1, NCOLS + 1):
        ws.cell(row=meta_row, column=c).fill = _fill("FFFFFF")

    # E2 — "Player:" right-justified
    player_label = ws.cell(row=meta_row, column=5, value="Player:")
    player_label.font = _font(bold=True, size=15)
    player_label.fill = _fill("FFFFFF")
    player_label.alignment = Alignment(horizontal="right", vertical="center")

    # F2:H2 — player name, merged, left-justified
    ws.merge_cells(start_row=meta_row, start_column=6, end_row=meta_row, end_column=8)
    player_val = ws.cell(row=meta_row, column=6, value=session_info.get("player", ""))
    player_val.font = _font(size=15)
    player_val.fill = _fill("FFFFFF")
    player_val.alignment = _left()

    # J2 — "Coach:" right-justified
    coach_label = ws.cell(row=meta_row, column=10, value="Coach:")
    coach_label.font = _font(bold=True, size=15)
    coach_label.fill = _fill("FFFFFF")
    coach_label.alignment = Alignment(horizontal="right", vertical="center")

    # K2:L2 — coach name, merged, left-justified
    ws.merge_cells(start_row=meta_row, start_column=11, end_row=meta_row, end_column=12)
    coach_val = ws.cell(row=meta_row, column=11, value=session_info.get("coach", ""))
    coach_val.font = _font(size=15)
    coach_val.fill = _fill("FFFFFF")
    coach_val.alignment = _left()

    ws.row_dimensions[meta_row].height = 20
    current_row += 1

    # ── Row 3: Date at E3, Week at J3 ────────────────────────────────────────
    week_row = current_row
    for c in range(1, NCOLS + 1):
        ws.cell(row=week_row, column=c).fill = _fill("FFFFFF")

    # E3 — "Date:" right-justified
    date_label_cell = ws.cell(row=week_row, column=5, value="Date:")
    date_label_cell.font = _font(bold=True, size=15)
    date_label_cell.fill = _fill("FFFFFF")
    date_label_cell.alignment = Alignment(horizontal="right", vertical="center")

    # F3:H3 — date value, merged, left-justified
    ws.merge_cells(start_row=week_row, start_column=6, end_row=week_row, end_column=8)
    date_val = ws.cell(row=week_row, column=6, value=session_info.get("date", ""))
    date_val.font = _font(size=15)
    date_val.fill = _fill("FFFFFF")
    date_val.alignment = _left()

    # J3 — "Week:" right-justified
    week_label_cell = ws.cell(row=week_row, column=10, value="Week:")
    week_label_cell.font = _font(bold=True, size=15)
    week_label_cell.fill = _fill("FFFFFF")
    week_label_cell.alignment = Alignment(horizontal="right", vertical="center")

    # K3:L3 — week value, merged, left-justified
    ws.merge_cells(start_row=week_row, start_column=11, end_row=week_row, end_column=12)
    week_val = ws.cell(row=week_row, column=11, value=session_info.get("week", ""))
    week_val.font = _font(size=15)
    week_val.fill = _fill("FFFFFF")
    week_val.alignment = _left()

    ws.row_dimensions[week_row].height = 20
    current_row += 2

    # ── Sections ─────────────────────────────────────────────────────────────
    all_stat_rows = []

    for section in sections:
        sec_name = section["section_name"]
        rows = section["rows"]

        # Empty spacer row + section title (no merge)
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        _merge_title(ws, current_row, sec_name, 1, NCOLS, C_SECTION_BG, C_SECTION_FG, size=15, bold=True, align="left")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers
        col_headers = [
            "Club", "Level", "Target Type", "Distance (yd)",
            "Target (raw)", "Actual (raw)", "Target %",
            "Attempts", "Successes", "Actual %", "Goal %",
            "Goal Status"
        ]
        _col_header_row(ws, current_row, col_headers, NCOLS)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        for i, row_data in enumerate(rows):
            club      = row_data["club"]
            level     = row_data["level"]
            ttype     = row_data["target_type"]
            dist      = row_data["distance_yd"]
            stats     = row_data["stats"]
            bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"

            # ── Populate cells ───────────────────────────────────────────
            cells_vals = {
                COL_CLUB:   club,
                COL_LEVEL:  level,
                COL_TTYPE:  ttype,
                COL_DIST:   dist,
                COL_ATT:    stats["attempts"] if stats["attempts"] else "",
                COL_SUCC:   stats["successes"] if stats["successes"] is not None else "",
            }

            # Static computed values from Python
            cells_vals[COL_TRAW]  = stats["target_raw"] or ""
            cells_vals[COL_ARAW]  = stats["actual_raw"] or ""
            cells_vals[COL_TPCT]  = stats["target_pct"]
            cells_vals[COL_APCT]  = stats["actual_pct"]
            cells_vals[COL_GOAL]  = stats["actual_pct"] / stats["target_pct"] if (stats["actual_pct"] is not None and stats["target_pct"]) else ""
            cells_vals[COL_STATUS] = stats["goal_status"] or ""

            # Write all cells
            for col_idx, val in cells_vals.items():
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.fill = _fill(bg)
                cell.alignment = _center()
                cell.font = _font(size=14)
                cell.border = FULL_BORDER

            # No special cell highlighting — all white

            tpct_cell = ws.cell(row=current_row, column=COL_TPCT)
            tpct_cell.number_format = "0%"

            apct_cell = ws.cell(row=current_row, column=COL_APCT)
            apct_cell.number_format = "0%"
            _pct_style(apct_cell, stats.get("actual_pct"), stats.get("target_pct"))

            goal_cell = ws.cell(row=current_row, column=COL_GOAL)
            goal_cell.number_format = "0%"
            _pct_style(goal_cell, stats.get("actual_pct"), stats.get("target_pct"))

            status_cell = ws.cell(row=current_row, column=COL_STATUS)
            _goal_status_style(status_cell, stats.get("goal_status"))
            status_cell.alignment = _center()

            all_stat_rows.append(current_row)
            current_row += 1



    # ── Overall summary ───────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 15
    current_row += 1
    _merge_title(ws, current_row, "Overall", 1, NCOLS, C_HEADER_BG, C_HEADER_FG, size=15, bold=True, align="left")
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    for col_idx, h in enumerate(["Attempts", "Successes", "Success %", "Goals", "Goals Met", "Goal %"], 1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = Font(bold=True, italic=True, size=13, name="Arial")
        cell.fill = _fill("F2F2F2")
        cell.alignment = _center()
        cell.border = FULL_BORDER
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    ov = overall
    ov_vals = {
        1: ov["total_attempts"],
        2: ov["total_successes"],
        3: f'{round(ov["success_pct"] * 100):.0f}%' if ov["success_pct"] is not None else "",
        4: ov["goals_total"],
        5: ov["goals_met"],
        6: f'{round(ov["goal_pct"] * 100):.0f}%' if ov["goal_pct"] is not None else "",
    }
    for col_idx, val in ov_vals.items():
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = _font(bold=False, size=15)
        cell.fill = _fill("FFFFFF")
        cell.alignment = _center()
        cell.border = FULL_BORDER
    # Color Success % (col 3) and Goal % (col 6)
    if ov["success_pct"] is not None:
        _pct_style(ws.cell(row=current_row, column=3), ov["success_pct"], 1.0, size=15)
    if ov["goal_pct"] is not None:
        _pct_style(ws.cell(row=current_row, column=6), ov["goal_pct"], 1.0, size=15)
    current_row += 1

    # Notes
    ws.row_dimensions[current_row].height = 15
    current_row += 1
    _merge_title(ws, current_row, "Additional Notes", 1, NCOLS, "FFFFFF", "000000", size=15, bold=True, align="left")
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row + 4, end_column=12
    )
    notes_cell = ws.cell(row=current_row, column=1)
    notes_cell.fill = _fill("F5F5F5")
    notes_cell.alignment = _left()
    ws.row_dimensions[current_row].height = 80


def build_raw_data_sheet(ws, df):
    ws.title = "Raw Data"
    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=14)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _center()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 2)

    import pandas as _pd
    if isinstance(df, _pd.DataFrame):
        rows_iter = df.to_dict("records")
    else:
        try:
            rows_iter = df.to_dicts()
        except AttributeError:
            rows_iter = df.to_dict("records")

    for row_idx, row in enumerate(rows_iter, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row[h]
            if val is not None and isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _font(size=13)
            if row_idx % 2 == 0:
                cell.fill = _fill(C_ALT_ROW)


def build_excel_report(
    df,
    session_info: dict,
    club_configs: list[dict],
    logo_path: str = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Raw Data sheet ──────────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw Data")
    build_raw_data_sheet(ws_raw, df)

    # ── Build section structure ─────────────────────────────────────────
    section_order = ["Putting", "Wedge Play", "Approach", "Driving", "Other"]
    sections_dict: dict[str, list] = {}

    for cfg in club_configs:
        sec = cfg.get("section", "Other")
        if sec not in sections_dict:
            sections_dict[sec] = []
        stats = compute_club_stats(
            df,
            club_name=cfg["club"],
            target_type=cfg["target_type"],
            distance_yd=cfg["distance_yd"],
            level=cfg["level"],
        )
        sections_dict[sec].append({**cfg, "stats": stats})

    sections = []
    for sec in section_order:
        if sec in sections_dict:
            sections.append({"section_name": sec, "rows": sections_dict[sec]})
    for sec in sections_dict:
        if sec not in section_order:
            sections.append({"section_name": sec, "rows": sections_dict[sec]})

    all_stats = [row["stats"] for sec in sections for row in sec["rows"]]
    overall = compute_overall(all_stats)

    # ── Report sheet ────────────────────────────────────────────────────
    ws_report = wb.create_sheet("Report")
    ws_report.sheet_view.showGridLines = False
    build_report_sheet(ws_report, session_info, sections, "Raw Data", overall, logo_path=logo_path)

    # Move Report to front
    wb.move_sheet("Report", offset=-wb.index(wb["Report"]))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
